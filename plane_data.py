from typing import Any, List, Tuple, Dict
from datetime import datetime, time
from openpyxl import load_workbook, Workbook

FILE_PATH = './DATASAMPLE.txt'
# CAT048UAP_PATH = './CAT048UAP.xlsx'


def read_data(path: str) -> List[str]:
    """读取数据"""
    with open(path, 'r', encoding='UTF-8') as f:
        return [line.strip() for line in f]


def trans_stap_to_time(stap: float) -> time:
    """将时间戳转换成time对象"""
    hour, minute, second, microsec = 0, 0, 0, 0
    hour = int(stap // 3600)
    stap -= 3600 * hour
    hour %= 24
    minute = int(stap // 60)
    stap -= 60 * minute
    second = int(stap // 1)
    stap -= 1 * second
    microsec = int(stap * 1e6)
    return time(hour=hour, minute=minute, second=second, microsecond=microsec)


# def load_UAP() -> List['CAT048UAPItem']:
#     """从Excel中读取UAP表"""
#     result = []
#     work_book = load_workbook(CAT048UAP_PATH, read_only=True, data_only=True)
#     sheet = work_book.active
#     for FRN, DataItem, DataItemDesc, Length in list(sheet.values)[1:]:
#         result.append(CAT048UAPItem(FRN, DataItem, DataItemDesc, Length))
#     return result


# class CAT048UAPItem:
#     """UAP表中的一项"""
#     def __init__(self, FRN: int, DataItem: str, DataItemDescription: str, Length: str) -> None:
#         self.FRN = FRN
#         self.DataItem = DataItem
#         self.DataItemDescription = DataItemDescription
#         self.Length = Length


class SecondaryRadar:
    """二次雷达数据 基类"""
    def __init__(self, source: str) -> None:
        soc = source.split()
        self.recv_time: datetime = self._read_utc_str(soc[0])   # 接收到数据的时间
        self.__source: str = soc[1]
        self.__analysis()
    
    @classmethod
    def get_cat(cls, source: str) -> int:
        """获取原数据的CAT"""
        return cls(source).CAT
    
    def __analysis(self) -> None:
        """分析数据"""
        self.HDLC_address = self._read_bits(1)  # HDLC地址字段
        self.HDLC_control = self._read_bits(1)  # HDLC控制字段
        self.CAT = self._read_bits(1)          # CAT数据种类
        self.LEN = self._read_bits(2)          # 数据帧的总长度
        self.FSPEC, _ = self._read_bits_bin_FX()   # UAP表的数据索引

    def _read_utc_str(self, utc: str) -> datetime:
        """解析时间"""
        day, month, year, hour, minute, second = 0, 0, 0, 0, 0, 0
        date_str, sec_str = utc.split(':')
        day = int(date_str[-2:])        # 日
        month = int(date_str[-4:-2])    # 月
        year = int(date_str[:-4])       # 年
        sec = float(sec_str)
        hour = int(sec // 3600)          # 时
        sec -= 3600 * hour
        minute = int(sec // 60)          # 分
        sec -= 60 * minute
        second = int(sec // 1)           # 秒
        sec -= 1 * second
        microsecond = int(sec * 1e6)
        return datetime(year=year, month=month, day=day,
                        hour=hour, minute=minute, second=second, microsecond=microsecond)

    def _read_bits(self, bits: int) -> int:
        """从数据中读取bits个比特"""
        soc = self._read_bits_str(bits)
        return int(soc, 16)

    def _read_bits_str(self, bits: int) -> str:
        """读取bits个比特，返回原始16进制串"""
        soc = self.__source[:bits*2]
        self.__source = self.__source[bits*2:]
        return soc
    
    def _read_bits_bin(self, bits: int) -> str:
        """读取bits个比特，返回01串"""
        length = bits * 8
        soce = self._read_bits(bits)
        result = bin(soce)[2:]
        return '0' * (length - len(result)) + result

    def _read_bits_FX(self, min: int = 1, step: int = 1) -> Tuple[int, int]:
        """
        依据FX位读取若干个比特
        至少读取min个
        一次步进step个比特
        :return 读取的数据，实际读取的比特数
        """
        soc, read_num = self._read_bits_str_FX(min, step)
        return int(soc, 16), read_num
    
    def _read_bits_str_FX(self, min: int = 1, step: int = 1) -> Tuple[str, int]:
        """
        依据FX位读取若干个比特
        至少读取min个
        一次步进step个
        :return 读取的数据的16进制串，实际读取的比特数
        """
        read_num = 0
        result_list = []
        result_list.append(self._read_bits_str(min))
        read_num += min
        while (int(result_list[-1], 16) & 1) != 0:
            result_list.append(self._read_bits_str(step))
            read_num += step
        soc = ''.join(result_list)
        return soc, read_num
    
    def _read_bits_bin_FX(self, min: int = 1, step: int = 1) -> Tuple[str, int]:
        """
        依据FX位读取若干个比特
        至少读取min个
        一次步进step个
        :return 读取的数据的01串，实际读取的比特数
        """
        soce, num = self._read_bits_FX(min, step)
        length = num * 8
        result = bin(soce)[2:]
        return '0' * (length - len(result)) + result, num
    
    def dump_json(self) -> Dict[str, Any]:
        pass


class SecondaryRadar048(SecondaryRadar):
    """二次雷达数据CAT048"""
    # CAT048UAP: List[CAT048UAPItem] = load_UAP()         # 完整的UAP表

    def __init__(self, source: str) -> None:
        super().__init__(source)
        self.__set_UAP_Item()
    
    def __set_UAP_Item(self) -> None:
        """根据FSPEC和UAP设置字段值"""
        # Data SOurce Identifier: SAC SIC
        if self.FSPEC[1-1] == '1':
            self.SAC = self._read_bits(1)      # 区域码
            self.SIC = self._read_bits(1)      # 雷达的编码
        # Time-of-Day
        if self.FSPEC[2-1] == '1':
            stap = self._read_bits(3) / 128
            self.data_time = trans_stap_to_time(stap)  # 数据包产生的时间
        # Target Report Descriptor
        if self.FSPEC[3-1] == '1':
            self._read_bits_str_FX()
        # Measured Position in Slant Polar Coordinates
        if self.FSPEC[4-1] == '1':
            self.polar_diameter = self._read_bits(2) / 256 * 1852     # 极径m
            self.polar_angle = self._read_bits(2) * 360 / (1<<16)      # 极角°
        # Mode-3/A Code in Octal Representation
        if self.FSPEC[5-1] == '1':
            self._read_bits_str(2)
        # Flight Level in Binary Representation
        if self.FSPEC[6-1] == '1':
            tap = self._read_bits(2)
            if tap < (1<<14):
                self.FL = tap / 4 * 30.48      # 飞行高度 m
        # Radar Plot Characteristics
        if self.FSPEC[7-1] == '1':
            rpc, _ = self._read_bits_bin_FX()
            for v in rpc:
                if v == '1':
                    self._read_bits_str(1)
        # Aircraft Address
        if self.FSPEC[8+1-1] == '1':
            self.ICAO = self._read_bits(3)  # 飞机的ICAO码
        
        if len(self.FSPEC) <= 8:
            return

        # Aircraft Identification
        if self.FSPEC[9+1-1] == '1':
            self._read_bits_str(6)
        # Mode S MB Data
        if self.FSPEC[10+1-1] == '1':
            MB_len = self._read_bits(1)
            self._read_bits_str(8 * MB_len)
        # Track Number
        if self.FSPEC[11+1-1] == '1':
            self.track_number = self._read_bits(2) # 航迹号
        # Calculatede Position in Cartesian Coordinates
        if self.FSPEC[12+1-1] == '1':
            self.position_X = self._read_bits(2) / 128 * 1852  # X坐标，单位m
            self.position_Y = self._read_bits(2) / 128 * 1852  # Y坐标，单位m
        # Calculated Track Velocity in Polar Representation
        if self.FSPEC[13+1-1] == '1':
            self.polar_track_velocity = self._read_bits(2) * 1852 / (1<<14)   # 极坐标的轨道速度 m/s
            self.polar_track_heading = self._read_bits(2) * 360 / (1<<16)      # 速度的方向 °
        # Track Status
        if self.FSPEC[14+1-1] == '1':
            self._read_bits_str_FX()

        if len(self.FSPEC) <= 16:
            return

        # Track QUality
        if self.FSPEC[15+2-1] == '1':
            self._read_bits_str(4)
        # Warning/Error Conditions
        if self.FSPEC[16+2-1] == '1':
            self._read_bits_str_FX()
        # Mode-3/A Code Confidence Indicator
        if self.FSPEC[17+2-1] == '1':
            self._read_bits_str(2)
        # Mode-C Code and Confidence Indicator
        if self.FSPEC[18+2-1] == '1':
            self._read_bits_str(4)
        # Height Measured by 3D Radar
        # if self.FSPEC[19+2-1] == '1':
        #     self.D3_height = self._read_bits(2) * 25 * 0.3 # 三维雷达测定的目标高度，使用平均海平面作为0基准面
    
    def dump_json(self) -> Dict[str, Any]:
        """获取对象的数据"""
        member = dir(self)
        return {
            "recv_time": self.recv_time if 'recv_time' in member else None,
            "HDLC_address": self.HDLC_address if 'HDLC_address' in member else None,
            "HDLC_control": self.HDLC_control if 'HDLC_control' in member else None,
            "CAT": self.CAT if 'CAT' in member else None,
            "LEN": self.LEN if 'LEN' in member else None,
            "FSPEC": self.FSPEC if 'FSPEC' in member else None,
            "SAC": self.SAC if 'SAC' in member else None,
            "SIC": self.SIC if 'SIC' in member else None,
            "Time-of-Day": self.data_time if 'data_time' in member else None,
            "polar diameter": self.polar_diameter if 'polar_diameter' in member else None,
            "polar angle": self.polar_angle if 'polar_angle' in member else None,
            "FL": self.FL if 'FL' in member else None,
            "ICAO": self.ICAO if 'ICAO' in member else None,
            "track number": self.track_number if 'track_number' in member else None,
            "position X": self.position_X if 'position_X' in member else None,
            "position Y": self.position_Y if 'position_Y' in member else None,
            "polar track velocity": self.polar_track_velocity if 'polar_track_velocity' in member else None,
            "polar track heading": self.polar_track_heading if 'polar_track_heading' in member else None,
            # "3D height": self.D3_height if 'D3_height' in member else None,
        }
    
    @classmethod
    def load_excel(cls, datas: List['SecondaryRadar048'], path: str) -> None:
        """把结果写入文件"""
        heads = [
            "recv_time",
            # "HDLC_address",
            # "HDLC_control",
            "CAT",
            # "LEN",
            # "FSPEC",
            "SAC",
            "SIC",
            "Time-of-Day",
            "polar diameter",
            "polar angle",
            "FL",
            "ICAO",
            "track number",
            "position X",
            "position Y",
            "polar track velocity",
            "polar track heading",
        ]
        wb = Workbook()
        ws = wb.active
        for i, title in enumerate(heads):
            ws.cell(row=1, column=i+1, value=title)
        for i, data in enumerate(datas):
            json_data = data.dump_json()
            for j, title in enumerate(heads):
                ws.cell(row=i+2, column=j+1, value=json_data[title])
        wb.save(path)  


class SecondaryRadar034(SecondaryRadar):
    """二次雷达数据CAT034"""
    def __init__(self, source: str) -> None:
        super().__init__(source)
        self.__set_UAP_Item()

    def __set_UAP_Item(self) -> None:
        """根据FSPEC和UAP设置字段值"""
        # Data Source Identifier
        if self.FSPEC[1-1] == '1':
            self.SAC = self._read_bits(1)      # 区域码
            self.SIC =self._read_bits(1)       # 雷达的编码
        # Message Type
        if self.FSPEC[2-1] == '1':
            self.message_type = self._read_bits(1) # 消息类型
        # Time-of-Day
        if self.FSPEC[3-1] == '1':
            stap = self._read_bits(3) / 128
            self.data_time = trans_stap_to_time(stap)   # 数据包产生的时间
        # Sector Number
        if self.FSPEC[4-1] == '1':
            self.sector_number = self._read_bits(1)    # 扇区号
        # Antenna Rotation Period
        # if self.FSPEC[5-1] == '1':
        #     self.antenna_rotation_period = self._read_bits(2) / 128    # 天线旋转周期 s
        # # System Configuration and Status
        # if self.FSPEC[6-1] == '1':
        #     rpc, _ = self._read_bits_bin_FX()
        #     for v in rpc:
        #         if v == '1':
        #             self._read_bits_str(1)
        # # System Processing Mode
        # if self.FSPEC[7-1] == '1':
        #     rpc, _ = self._read_bits_bin_FX()
        #     for v in rpc:
        #         if v == '1':
        #             self._read_bits_str(1)
        
        # if len(self.FSPEC) <= 8:
        #     return
        
        # # Message Count Values
        # if self.FSPEC[8+1-1] == '1':
        #     rep = self._read_bits(1)
        #     self._read_bits_str(2 * rep)
        # # Generic Polar Window
        # if self.FSPEC[9+1-1] == '1':
        #     self._read_bits_str(8)
        # # Data Filter
        # if self.FSPEC[10+1-1] == '1':
        #     self._read_bits_str(1)
        # # 3D-Position of Date Source
        # if self.FSPEC[11+1-1] == '1':
        #     self._read_bits_str(8)
        # # Collimation Error
        # if self.FSPEC[12+1-1] == '1':
        #     self._read_bits_str(2)

    def dump_json(self) -> Dict[str, Any]:
        """获取对象的数据"""
        member = dir(self)
        return {
            "recv_time": self.recv_time if 'recv_time' in member else None,
            "HDLC_address": self.HDLC_address if 'HDLC_address' in member else None,
            "HDLC_control": self.HDLC_control if 'HDLC_control' in member else None,
            "CAT": self.CAT if 'CAT' in member else None,
            "LEN": self.LEN if 'LEN' in member else None,
            "FSPEC": self.FSPEC if 'FSPEC' in member else None,
            "SAC": self.SAC if 'SAC' in member else None,
            "SIC": self.SIC if 'SIC' in member else None,
            "Time-of-Day": self.data_time if 'data_time' in member else None,
            "Message Type": self.message_type if 'message_type' in member else None,
            "Sector Number": self.sector_number if 'sector_number' in member else None,
            # "Antenna Rotation Period": self.antenna_rotation_period if 'antenna_rotation_period' in member else None,
        }

    @classmethod
    def load_excel(cls, datas: List['SecondaryRadar034'], path: str) -> None:
        """把结果写入文件"""
        heads = [
            "recv_time",
            "HDLC_address",
            "HDLC_control",
            "CAT",
            "LEN",
            "FSPEC",
            "SAC",
            "SIC",
            "Time-of-Day",
            "Message Type",
            "Sector Number",
            # "Antenna Rotation Period",
        ]
        wb = Workbook()
        ws = wb.active
        for i, title in enumerate(heads):
            ws.cell(row=1, column=i+1, value=title)
        for i, data in enumerate(datas):
            json_data = data.dump_json()
            for j, title in enumerate(heads):
                ws.cell(row=i+2, column=j+1, value=json_data[title])
        wb.save(path)  

if __name__ == '__main__':
    data_list_048 = []
    # data_list_034 = []
    for data in read_data(FILE_PATH):
        cat = SecondaryRadar.get_cat(data)
        if cat == 48:
            a = SecondaryRadar048(data)
            data_list_048.append(a)
        # elif cat == 34:
        #     a = SecondaryRadar034(data)
        #     data_list_034.append(a)
        # else:
        #     print('??????')
    # SecondaryRadar034.load_excel(data_list_034, 'test034.xlsx')
    SecondaryRadar048.load_excel(data_list_048, 'test048.xlsx')
