from typing import Any, Generator, List
import re
from datetime import datetime
from openpyxl import Workbook

FILE_PATH = './AWOS202008010000.SJN'

class AutomaticObservationData:
    """民航自动观测数据"""
    def __init__(self) -> None:
        self.date_time : datetime = datetime.now()  # 时间组
        self.runway_num : str = None    # 跑道编号
        self.TDZ_num : str = None       # 着陆端编号
        self.MID_num : str = None       # 中间端标识
        self.END_num : str = None       # 停止端编号
        self.TDZ_SingleData : SingleEndedData = None      # TDZ端自观数据
        self.MID_SingleData : SingleEndedData = None      # MID端自观数据
        self.END_SingleData : SingleEndedData = None      # END端自观数据
    
    def keys(self) -> List[str]:
        return [
            'date_time',
            'runway_num',
            'TDZ_num',
            'MID_num',
            'END_num',
            'TDZ_SingleData',
            'MID_SingleData',
            'END_SingleData'
        ]
    
    def __getitem__(self, key: str) -> Any:
        return getattr(self, key)
    
    @classmethod
    def fromSource(cls, source: str) -> 'AutomaticObservationData':
        """解析数据，返回对象"""
        result = cls()
        sourc_list = source.splitlines()
        date_value = [int(v) for v in sourc_list[0].split(':')]
        result.date_time = datetime(year=date_value[0], month=date_value[1], day=date_value[2], hour=date_value[3], minute=date_value[4])
        runway_value = sourc_list[1].split()
        result.runway_num = runway_value[0]
        result.TDZ_num = runway_value[1]
        result.MID_num = runway_value[2]
        result.END_num = runway_value[3]
        result.TDZ_SingleData = SingleEndedData.fromSource(sourc_list[2])
        result.MID_SingleData = SingleEndedData.fromSource(sourc_list[3])
        result.END_SingleData = SingleEndedData.fromSource(sourc_list[4])
        return result
    
    @classmethod
    def to_excel(cls, data_list: List['AutomaticObservationData'], path: str) -> None:
        """导出Excel"""
        wb = Workbook()
        ws = wb.active
        title_list = ['date_time', 'runway_num', 'TDZ_num', 'MID_num', 'END_num', 
            'TDZ_DATA_num', 'TDZ_DATA_RVR_1', 'TDZ_DATA_RVR_10', 'TDZ_DATA_MOR_1', 'TDZ_DATA_MOR_10', 'TDZ_DATA_B1', 'TDZ_DATA_windS2', 'TDZ_DATA_windF2', 'TDZ_DATA_winS10', 'TDZ_DATA_windF10', 'TDZ_DATA_windS', 'TDZ_DATA_windF', 'TDZ_DATA_Qnh', 'TDZ_DATA_Qfe', 'TDZ_DATA_Temp', 'TDZ_DATA_Hum', 'TDZ_DATA_Td', 'TDZ_DATA_roadTemp', 'TDZ_DATA_LowCBase', 'TDZ_DATA_MediaCBase', 'TDZ_DATA_HighCBase', 'TDZ_DATA_VV', 'TDZ_DATA_WEA', 'TDZ_DATA_Pi', 'TDZ_DATA_PREC', 
            'MID_DATA_num', 'MID_DATA_RVR_1', 'MID_DATA_RVR_10', 'MID_DATA_MOR_1', 'MID_DATA_MOR_10', 'MID_DATA_B1', 'MID_DATA_windS2', 'MID_DATA_windF2', 'MID_DATA_winS10', 'MID_DATA_windF10', 'MID_DATA_windS', 'MID_DATA_windF', 'MID_DATA_Qnh', 'MID_DATA_Qfe', 'MID_DATA_Temp', 'MID_DATA_Hum', 'MID_DATA_Td', 'MID_DATA_roadTemp', 'MID_DATA_LowCBase', 'MID_DATA_MediaCBase', 'MID_DATA_HighCBase', 'MID_DATA_VV', 'MID_DATA_WEA', 'MID_DATA_Pi', 'MID_DATA_PREC', 
            'END_DATA_num', 'END_DATA_RVR_1', 'END_DATA_RVR_10', 'END_DATA_MOR_1', 'END_DATA_MOR_10', 'END_DATA_B1', 'END_DATA_windS2', 'END_DATA_windF2', 'END_DATA_winS10', 'END_DATA_windF10', 'END_DATA_windS', 'END_DATA_windF', 'END_DATA_Qnh', 'END_DATA_Qfe', 'END_DATA_Temp', 'END_DATA_Hum', 'END_DATA_Td', 'END_DATA_roadTemp', 'END_DATA_LowCBase', 'END_DATA_MediaCBase', 'END_DATA_HighCBase', 'END_DATA_VV', 'END_DATA_WEA', 'END_DATA_Pi', 'END_DATA_PREC']
        for i, title in enumerate(title_list):
            ws.cell(row=1, column=i+1, value=title)
        for i, data in enumerate(data_list):
            for j, title in enumerate(title_list):
                if j < 5:
                    value = data[title]
                else:
                    index = title[:3] + '_SingleData'
                    value = data[index][title[9:]]
                ws.cell(row=i+2, column=j+1, value=value)
        wb.save(path)

class SingleEndedData:
    """单端的自观数据"""
    def __init__(self) -> None:
        self.num : str = None           # 端编号
        self.RVR_1 : str = None         # 1分钟RVR平均
        self.RVR_10 : str = None        # 10分钟平均
        self.MOR_1 : str = None         # 1分钟MOR平均
        self.MOR_10 : str = None        # 10分钟MOR平均
        self.B1 : str = None            # 背景亮度一分钟平均值
        self.windS2 : str = None        # 2分钟平均风速
        self.windF2 : str = None        # 2分钟平均风向
        self.winS10 : str = None        # 10分钟平均风速
        self.windF10 : str = None       # 10分钟平均风向
        self.windS : str = None         # 最大阵风风速
        self.windF : str = None         # 最大阵风风向
        self.Qnh : str = None           # 修正海压
        self.Qfe : str = None           # 场压
        self.Temp : str = None          # 温度
        self.Hum : str = None           # 相对湿度
        self.Td : str = None            # 露点温度
        self.roadTemp : str = None      # 道面温度
        self.LowCBase : str = None      # 低云层高度
        self.MediaCBase : str = None    # 中云层高度
        self.HighCBase : str = None     # 高云层高度
        self.VV : str = None            # 垂直能见度
        self.WEA : str = None           # 天气现象
        self.Pi : str = None            # 降水强度
        self.PREC : str = None          # 降水量
    
    def keys(self) -> List[str]:
        return [
            'self.num',
            'self.RVR_1',
            'self.RVR_10',
            'self.MOR_1',
            'self.MOR_10',
            'self.B1',
            'self.windS2',
            'self.windF2',
            'self.winS10',
            'self.windF10',
            'self.windS',
            'self.windF',
            'self.Qnh',
            'self.Qfe',
            'self.Temp',
            'self.Hum',
            'self.Td',
            'self.roadTemp',
            'self.LowCBase',
            'self.MediaCBase',
            'self.HighCBase',
            'self.VV',
            'self.WEA',
            'self.Pi',
            'self.PREC'
        ]
    
    def __getitem__(self, key: str) -> Any:
        return getattr(self, key)
    
    @classmethod
    def fromSource(cls, source: str) -> 'SingleEndedData':
        result = cls()
        source_list = source.split()
        result.num = source_list[0] if source_list[0] != '///' else None
        result.RVR_1 = source_list[1] if source_list[1] != '///' else None
        result.RVR_10 = source_list[2] if source_list[2] != '///' else None
        result.MOR_1 = source_list[3] if source_list[3] != '///' else None
        result.MOR_10 = source_list[4] if source_list[4] != '///' else None
        result.B1 = source_list[5] if source_list[5] != '///' else None
        result.windS2 = source_list[6] if source_list[6] != '///' else None
        result.windF2 = source_list[7] if source_list[7] != '///' else None
        result.winS10 = source_list[8] if source_list[8] != '///' else None
        result.windF10 = source_list[9] if source_list[9] != '///' else None
        result.windS = source_list[10] if source_list[10] != '///' else None
        result.windF = source_list[11] if source_list[11] != '///' else None
        result.Qnh = source_list[12] if source_list[12] != '///' else None
        result.Qfe = source_list[13] if source_list[13] != '///' else None
        result.Temp = source_list[14] if source_list[14] != '///' else None
        result.Hum = source_list[15] if source_list[15] != '///' else None
        result.Td = source_list[16] if source_list[16] != '///' else None
        result.roadTemp = source_list[17] if source_list[17] != '///' else None
        result.LowCBase = source_list[18] if source_list[18] != '///' else None
        result.MediaCBase = source_list[19] if source_list[19] != '///' else None
        result.HighCBase = source_list[20] if source_list[20] != '///' else None
        result.VV = source_list[21] if source_list[21] != '///' else None
        result.WEA = source_list[22] if source_list[22] != '///' else None
        result.Pi = source_list[23] if source_list[23] != '///' else None
        result.PREC = source_list[24] if source_list[24] != '///' else None
        return result


def read_data() -> str:
    """读取文件数据"""
    with open(FILE_PATH, 'r', encoding='UTF-8') as f:
        return f.read()

def analysis_data(source: str) -> Generator:
    """拆分文件数据为单个数据项"""
    pattern = r'ZCZC\n.*?\nNNNN'
    for value in re.findall(pattern, source, re.M | re.S):
        yield value[5:-5]


if __name__ == '__main__':
    source = read_data()
    data_list = analysis_data(source)
    su_list = []
    for data in data_list:
        a = AutomaticObservationData.fromSource(data)
        su_list.append(a)
    AutomaticObservationData.to_excel(su_list, '自观测.xlsx')
